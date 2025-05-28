import requests, os, datetime, sys
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk
from threading import Thread

def resource_path(relative_path):
    """ Retorna o caminho absoluto para o recurso, necessário para a execução do PyInstaller """
    try:
        # PyInstaller cria um diretório temporário e armazena o caminho em _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def fetch_data(url):
    response = requests.get(url)
    response.raise_for_status()
    return response.text

def parse_html(html):
    soup = BeautifulSoup(html, 'html.parser')
    traffic_info = []
    updates = soup.find_all('div', class_='status-list')[0].find_all('span', class_='name-route')
    
    for update in updates:
        data_status = update.find_next_sibling('div')
        details = data_status.find_all('div')[-1]
        route = update.text.strip()
        status = details.find('span').text.strip()
        last_updated = details.find('em').text.strip()
        traffic_info.append({"Route": route, "Status": status, "Last Updated": last_updated})
    
    return traffic_info

def create_excel(data):
    df = pd.DataFrame(data)
    wb = Workbook()
    ws = wb.active
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    folder_name = "Arquivos_Exel"
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_time_%H-%M-%S")
    filename = f"{folder_name}/{timestamp}-Alertas_de_Trafego.xlsx"
    wb.save(filename)

def show_loading_screen():
    loading_window = tk.Tk()
    loading_window.title("Carregando")
    loading_window.geometry('300x100+500+300')

    # Usar a função resource_path para localizar o ícone
    icon_path = resource_path('assets/Logo.ico')
    loading_window.iconbitmap(icon_path)

    tk.Label(loading_window, text="Aguarde 5s...\nGerando arquivo excel...", font=('Helvetica', 14)).pack(expand=True, pady=30)
    loading_window.after(5000, loading_window.destroy)  # Automatically close after 5 seconds
    loading_window.mainloop()
def main():
    thread = Thread(target=show_loading_screen)
    thread.start()
    url = "https://csg.com.br"
    html_content = fetch_data(url)
    traffic_data = parse_html(html_content)
    create_excel(traffic_data)
    thread.join()

if __name__ == "__main__":
    main()
